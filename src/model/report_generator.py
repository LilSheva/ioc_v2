from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

class ReportGenerator:
    def generate_query_data(self, found_iocs, ioc_config):
        query_groups = {}
        ordered_ioc_names = ioc_config.keys()

        for ioc_type in ordered_ioc_names:
            indicators = found_iocs.get(ioc_type, [])
            if not indicators: continue
            
            query_groups[ioc_type] = []
            templates = ioc_config.get(ioc_type, {}).get("query_templates", {})
            if not templates: continue
                
            cleaned_list = sorted([ind.replace('[.]', '.') for ind in indicators])
            for system, query_list in templates.items():
                if not query_list: continue
                joiner = " OR " if system == "MP10" else " || "
                for template in query_list:
                    parts = [template.format(ioc=ioc) for ioc in cleaned_list]
                    query_groups[ioc_type].append({"system": system, "query": joiner.join(parts)})
        return query_groups

    def create_query_file(self, filepath, query_data, ioc_config):
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                for name in ioc_config.keys():
                    items = query_data.get(name, [])
                    if not items: continue
                    f.write(f"--- {{{name}}} ---\n\n")
                    system_groups = {}
                    for item in items:
                        if item['system'] not in system_groups: system_groups[item['system']] = []
                        system_groups[item['system']].append(item['query'])
                    for system, queries in system_groups.items():
                        f.write(f"Для {system}\n")
                        for query in queries: f.write(f"{query}\n")
                        f.write("\n")
            return True, f"Файл с запросами успешно сохранен: {filepath}"
        except Exception as e:
            return False, f"Ошибка при сохранении файла с запросами: {e}"

    def create_xlsx_report(self, filepath, found_iocs, ioc_config):
        try:
            wb = Workbook(); ws = wb.active; ws.title = "Индикаторы компрометации"
            grey_fill=PatternFill(start_color="D3D3D3",end_color="D3D3D3",fill_type="solid"); bold_font=Font(bold=True)
            thin_border_side=Side(border_style="thin",color="000000"); thin_border=Border(left=thin_border_side,right=thin_border_side,top=thin_border_side,bottom=thin_border_side)
            wrap_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            headers = ["№", "Дата\nОтчёта", "Статус\nАктивности\nNTA", "Статус\nАктивности\nSIEM (Tools)", "Статус\nАктивности\nSIEM (MP)", "Тип\nИндикатора", "Индикатор", "IOC", "Бюллетень", "Тип события"]
            ws.append(headers); ws.row_dimensions[1].height = 60
            
            row_counter = 1
            for ioc_type in ioc_config.keys():
                indicators = found_iocs.get(ioc_type, [])
                if not indicators: continue
                template = ioc_config.get(ioc_type, {}).get("report_template", {})
                for indicator in sorted(indicators):
                    ws.append([row_counter, "", template.get("Статус Активности NTA", ""), "---------------", template.get("Статус Активности SIEM (MP)", ""), template.get("Тип Индикатора", ioc_type), indicator, indicator.replace('[.]', '.'), "", "Фишинговая рассылка электронной почты. Вредоносные вложения"])
                    row_counter += 1
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                for cell in row: cell.border = thin_border
            for cell in ws[1]: cell.fill = grey_fill; cell.font = bold_font; cell.alignment = wrap_alignment
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=1); cell.fill = grey_fill; cell.font = bold_font; cell.alignment = Alignment(horizontal='center', vertical='center')
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    if cell.value:
                        lines = str(cell.value).split('\n')
                        cell_max_len = max(len(line) for line in lines)
                        if cell_max_len > max_length: max_length = cell_max_len
                ws.column_dimensions[col[0].column_letter].width = max_length + 5
            
            wb.save(filepath)
            return True, f"Отчет Excel успешно сохранен: {filepath}"
        except Exception as e:
            return False, f"Ошибка при сохранении отчета Excel: {e}"